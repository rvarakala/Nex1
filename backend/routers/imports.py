"""Bulk patient import — CSV upload for new clinics migrating from another system.

v1 scope:
  * Patients only (demographics + contact + light triage).
  * CSV files only.
  * Preserves existing MRDs from the source system if provided; otherwise
    auto-generates one in the clinic's normal sequence.
  * Skips rows whose mobile or MRD already exists in the target clinic.
  * Two-step preview/commit so the operator sees a tally + per-row diagnosis
    before any writes happen.

Routes (all require clinic_owner / super_admin):
  GET  /api/imports/patients/template          — Download CSV template
  POST /api/imports/patients/preview           — Validate uploaded CSV
  POST /api/imports/patients/commit            — Persist preview by import_id
"""
import csv
import io
import re
import uuid
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from auth import require_roles
from database import get_db
from models import Patient
from utils.serde import serialize_datetime


router = APIRouter(prefix="/api/imports", tags=["imports"])


# Canonical column order used by the template + parser. Aliases below let the
# parser accept common variants without forcing the operator to rename headers.
TEMPLATE_HEADERS = [
    "name", "age", "gender", "mobile", "existing_mrd",
    "dob", "email", "alternate_mobile", "address", "city",
    "state", "pincode", "occupation", "chief_complaint",
    "referral_source", "notes",
    # Visit-level fields (rich-import flow): one row = one visit.
    "visit_date", "bill_no", "tests", "diagnosis", "amount", "referring_doctor",
]

HEADER_ALIASES = {
    # Patient core
    "patient_name": "name", "full_name": "name", "pt_name": "name", "pt._name": "name", "pt_._name": "name",
    "phone": "mobile", "phone_number": "mobile", "mobile_number": "mobile",
    "ph_no": "mobile", "ph._no": "mobile", "ph_._no": "mobile",
    "sex": "gender",
    "mrd": "existing_mrd", "mrn": "existing_mrd", "old_mrd": "existing_mrd", "patient_id": "existing_mrd",
    "mr_no": "existing_mrd", "mr._no": "existing_mrd", "mr_._no": "existing_mrd",
    "date_of_birth": "dob", "birth_date": "dob",
    "alt_mobile": "alternate_mobile", "secondary_phone": "alternate_mobile",
    "address1": "address", "street_address": "address", "area": "address",
    "zip": "pincode", "zipcode": "pincode", "postal_code": "pincode",
    "complaint": "chief_complaint",
    "source": "referral_source", "lead_source": "referral_source",
    "remarks": "notes",
    # Visit / billing
    "date": "visit_date", "visit": "visit_date", "appointment_date": "visit_date",
    "bill": "bill_no", "bill_number": "bill_no", "invoice_no": "bill_no", "invoice_number": "bill_no",
    "bill._no": "bill_no", "bill_._no": "bill_no",
    "test": "tests", "tests_performed": "tests", "test_performed": "tests", "investigation": "tests", "procedure": "tests",
    "dx": "diagnosis", "impression": "diagnosis", "findings": "diagnosis",
    "amt": "amount", "fee": "amount", "charges": "amount", "total_paid": "amount", "paid_amount": "amount",
    "ref_dr": "referring_doctor", "ref._dr": "referring_doctor", "ref_._dr": "referring_doctor",
    "referring_dr": "referring_doctor", "referring_physician": "referring_doctor",
    "referral_doctor": "referring_doctor", "doctor": "referring_doctor", "physician": "referring_doctor",
    # Index column we just drop
    "s.no": "_drop", "s_no": "_drop", "s._no": "_drop", "sl_no": "_drop", "sno": "_drop",
}

GENDER_MAP = {
    "m": "Male", "male": "Male",
    "f": "Female", "female": "Female",
    "o": "Other", "other": "Other", "third": "Other",
}

MAX_ROWS = 5000  # Hard ceiling per upload — protects the API from a 1M-row dump.
PREVIEW_TTL_HOURS = 2  # Stored preview blobs are pruned after this.


# ---------- helpers --------------------------------------------------------

def _normalise_header(h: str) -> str:
    h = (h or "").strip().lower()
    # Replace common separators with underscore so "Pt.Name", "Pt Name", "Pt-Name"
    # and "Pt_Name" all collapse to "pt_name".
    for ch in (" ", "-", ".", "/"):
        h = h.replace(ch, "_")
    # Collapse multiple underscores
    while "__" in h:
        h = h.replace("__", "_")
    h = h.strip("_")
    return HEADER_ALIASES.get(h, h)


def _parse_age(value: str, dob: Optional[str]) -> Optional[int]:
    if value:
        try:
            n = int(float(str(value).strip()))
            if 0 <= n <= 130:
                return n
        except (ValueError, TypeError):
            pass
    if dob:
        try:
            d = datetime.fromisoformat(dob)
            yrs = (datetime.utcnow() - d).days // 365
            if 0 <= yrs <= 130:
                return yrs
        except (ValueError, TypeError):
            pass
    return None


def _parse_gender(value: str) -> Optional[str]:
    v = (value or "").strip().lower()
    return GENDER_MAP.get(v)


_MOBILE_RE = re.compile(r"\D+")


def _normalise_mobile(value: str) -> Optional[str]:
    if not value:
        return None
    digits = _MOBILE_RE.sub("", str(value))
    # Strip leading country code 91 if 12 digits (`919812345678` → `9812345678`).
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    if len(digits) == 10:
        return digits
    if 7 <= len(digits) <= 15:  # Permissive — international landlines etc.
        return digits
    return None


_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _validate_email(value: str) -> Optional[str]:
    v = (value or "").strip()
    return v if v and _EMAIL_RE.match(v) else None


def _parse_dob(value: str) -> Optional[str]:
    v = (value or "").strip()
    if not v:
        return None
    # Accept ISO yyyy-mm-dd, dd-mm-yyyy, dd/mm/yyyy.
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(v, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _parse_amount(value) -> float:
    """Permissive ₹ / float / "1,250.00" parser. Negatives / NaN → 0."""
    if value is None or value == "":
        return 0.0
    try:
        s = str(value).strip().replace(",", "").replace("\u20b9", "").replace("Rs.", "").replace("INR", "")
        s = re.sub(r"[^\d.\-]", "", s)
        if not s or s in ("-", "."):
            return 0.0
        n = float(s)
        return max(0.0, n)
    except (ValueError, TypeError):
        return 0.0


def _split_tests(raw: str) -> list[str]:
    """Tests column may be 'PTA+IMP', 'PTA, IMP, VEMP', or just 'PTA'.
    Returns a list of clean uppercase tokens (e.g. ['PTA', 'IMP', 'VEMP'])."""
    if not raw:
        return []
    s = str(raw).replace(",", "+").replace("/", "+").replace("&", "+").replace(";", "+")
    return [t.strip().upper() for t in s.split("+") if t.strip()]


async def _next_mrd(db, clinic_id: str, mrd_prefix: str) -> str:
    """Mirror of patients.py — same counter, so generated MRDs slot into the
    clinic's existing sequence."""
    now = datetime.utcnow()
    counter = await db.counters.find_one_and_update(
        {"_id": f"mrd:{clinic_id}:{now.year}"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    seq = counter["seq"] if counter else 1
    return f"{mrd_prefix}-{now.year}-{seq:06d}"


# ---------- template download ----------------------------------------------

@router.get("/patients/template")
async def download_patients_template(
    user=Depends(require_roles("clinic_owner", "super_admin")),
):
    """Returns a CSV with header row + 2 example rows so the operator can fill it in."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(TEMPLATE_HEADERS)
    writer.writerow([
        "Asha Iyer", "62", "Female", "9876543210", "OLD-1024",
        "1962-04-12", "asha@example.com", "", "12 Marine Drive", "Mumbai",
        "Maharashtra", "400001", "Retired Teacher", "Reduced hearing both ears",
        "Walk-in", "Long-term patient since 2019",
        "01-04-2026", "BILL-A-001", "PTA+IMP", "Bil. Mild SNHL", "2500", "Internal Medicine",
    ])
    writer.writerow([
        "Rahul Singh", "34", "Male", "9123456780", "",
        "", "", "", "", "Bengaluru",
        "Karnataka", "560001", "Software Engineer", "",
        "Doctor", "",
        "02-04-2026", "", "PTA", "Mild HF SNHL", "1500", "Dr. Mehta",
    ])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="audinexa_patients_template.csv"'},
    )


# ---------- preview --------------------------------------------------------

@router.post("/patients/preview")
async def preview_patients(
    file: UploadFile = File(...),
    user=Depends(require_roles("clinic_owner", "super_admin")),
    db=Depends(get_db),
):
    """Parses + validates the uploaded CSV against the current clinic's data.

    Returns:
      {
        import_id: "...",
        tally: { total, will_create, will_skip, will_fail },
        rows: [ { row_num, name, mobile, mrd, status, errors:[..] }, ... ],
        expires_at: iso,
      }

    No data is written to `patients` yet — the parsed payload is stashed in
    `import_jobs` keyed by import_id, ready to be committed in a second call.
    """
    fname = (file.filename or "").lower()
    is_xlsx = fname.endswith(".xlsx")
    is_csv = fname.endswith(".csv")
    if not (is_csv or is_xlsx):
        raise HTTPException(400, "Please upload a .csv or .xlsx file.")

    raw = await file.read()
    if len(raw) > 5 * 1024 * 1024:
        raise HTTPException(400, "File too large — please split into batches of <5MB.")

    if is_xlsx:
        # Excel — read the first sheet, treat row 1 as headers.
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise HTTPException(500, "Excel support not installed (openpyxl missing on server).")
        try:
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(400, f"Could not read Excel file: {exc}")
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None) or []
        raw_headers = [str(h).strip() if h is not None else "" for h in header_row]
        # Drop trailing empty headers (typical when Excel has phantom columns)
        while raw_headers and not raw_headers[-1]:
            raw_headers.pop()
        if not raw_headers:
            raise HTTPException(400, "Excel file is empty or missing a header row.")

        def _row_iter():
            for row in rows_iter:
                # Skip totally-empty rows (Excel often pads to 1048576).
                if all((c is None or str(c).strip() == "") for c in row):
                    continue
                # Convert each cell to a string, preserving DD-MM-YYYY etc.
                cleaned = []
                for c in row[:len(raw_headers)]:
                    if c is None:
                        cleaned.append("")
                    elif isinstance(c, datetime):
                        cleaned.append(c.strftime("%Y-%m-%d"))
                    elif isinstance(c, (int, float)):
                        # int/float — drop trailing .0 on whole numbers (mobiles, MRDs)
                        if isinstance(c, float) and c.is_integer():
                            cleaned.append(str(int(c)))
                        else:
                            cleaned.append(str(c))
                    else:
                        cleaned.append(str(c).strip())
                yield dict(zip(raw_headers, cleaned))
        reader_iter = _row_iter()
    else:
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = raw.decode("latin-1")
            except UnicodeDecodeError:
                raise HTTPException(400, "Could not decode file — please save as UTF-8 CSV.")

        reader = csv.DictReader(io.StringIO(text))
        raw_headers = reader.fieldnames or []
        if not raw_headers:
            raise HTTPException(400, "CSV is empty or missing a header row.")
        reader_iter = reader

    # Map raw headers to canonical names, keep track of which canonical fields
    # were present so per-row lookups are O(1).
    header_map = {h: _normalise_header(h) for h in raw_headers}
    canonical_present = set(header_map.values())
    missing_required = [c for c in ("name",) if c not in canonical_present]
    if missing_required:
        raise HTTPException(
            400,
            f"Missing required column(s): {', '.join(missing_required)}. "
            f"Download the template for the expected layout.",
        )

    # Existing data for duplicate detection — pulled in one query each.
    existing_mobiles: set[str] = set()
    existing_mrds: set[str] = set()
    async for doc in db.patients.find(
        {"clinic_id": user["clinic_id"]},
        {"_id": 0, "mobile": 1, "mrd": 1},
    ):
        if doc.get("mobile"):
            m = _normalise_mobile(doc["mobile"])
            if m:
                existing_mobiles.add(m)
        if doc.get("mrd"):
            existing_mrds.add(str(doc["mrd"]).strip().upper())

    rows_out: list[dict] = []
    parsed_for_commit: list[dict] = []
    seen_mobiles_in_file: set[str] = set()
    seen_mrds_in_file: set[str] = set()
    seen_visit_keys: set[str] = set()  # NEW: detects same-patient-same-day-same-bill TRUE dupes
    counts = {"will_create": 0, "will_skip": 0, "will_fail": 0}

    for idx, raw_row in enumerate(reader_iter, start=2):  # Start at 2 — row 1 is header.
        if idx - 1 > MAX_ROWS:
            raise HTTPException(
                400,
                f"Files larger than {MAX_ROWS} rows aren't supported in a single upload — please split.",
            )

        # Re-key row using canonical headers, drop columns aliased to "_drop" (e.g. S.NO).
        row = {header_map.get(k, k): (v.strip() if isinstance(v, str) else v) for k, v in raw_row.items()}
        row.pop("_drop", None)
        # Treat fully-blank lines as a soft EOF (Excel often appends them).
        if not any((v or "").strip() for v in row.values() if isinstance(v, str)):
            continue

        errors: list[str] = []
        name = (row.get("name") or "").strip()
        if not name:
            errors.append("Name is required")

        gender = _parse_gender(row.get("gender", ""))
        # Gender is now optional — many import sources don't capture it.
        if not gender:
            gender = None  # stored as None on the patient

        dob = _parse_dob(row.get("dob", ""))
        age = _parse_age(row.get("age", ""), dob)
        if age is None:
            errors.append("Age (or DOB in YYYY-MM-DD / DD-MM-YYYY) is required")

        mobile = _normalise_mobile(row.get("mobile", ""))
        email = _validate_email(row.get("email", ""))
        if not mobile:
            errors.append("Mobile / phone number is required")

        existing_mrd = (row.get("existing_mrd") or "").strip().upper() or None

        # Visit-level fields (rich import — one row = one visit).
        visit_date = _parse_dob(row.get("visit_date", ""))   # reuses date parser
        bill_no = (row.get("bill_no") or "").strip() or None
        tests_raw = (row.get("tests") or "").strip()
        tests_list = _split_tests(tests_raw)
        diagnosis = (row.get("diagnosis") or "").strip() or None
        amount = _parse_amount(row.get("amount"))
        ref_dr = (row.get("referring_doctor") or "").strip() or None

        # Duplicate detection — within file + against DB.
        # NEW SEMANTICS: a repeating MR.NO/mobile is treated as a FOLLOW-UP visit,
        # not a duplicate, **unless** the same (mrd OR mobile) + visit_date +
        # bill_no triplet is already present in this file (true duplicate row).
        is_followup = False
        dup_reason = None
        if existing_mrd:
            if existing_mrd in existing_mrds or existing_mrd in seen_mrds_in_file:
                is_followup = True
        if mobile and not is_followup:
            if mobile in existing_mobiles or mobile in seen_mobiles_in_file:
                is_followup = True
        # In-file true-duplicate guard: same patient + same date + same bill_no
        in_file_key = None
        if visit_date and bill_no and (existing_mrd or mobile):
            in_file_key = f"{existing_mrd or mobile}|{visit_date}|{bill_no}"
            if in_file_key in seen_visit_keys:
                dup_reason = f"Same patient + {visit_date} + bill {bill_no} already in this file"

        if errors:
            status = "fail"
            counts["will_fail"] += 1
        elif dup_reason:
            status = "skip"
            counts["will_skip"] += 1
            errors = [dup_reason]
        else:
            status = "followup" if is_followup else "ok"
            counts["will_create"] += 1
            if mobile:
                seen_mobiles_in_file.add(mobile)
            if existing_mrd:
                seen_mrds_in_file.add(existing_mrd)
            if in_file_key:
                seen_visit_keys.add(in_file_key)

        rows_out.append({
            "row_num": idx,
            "name": name or "(missing)",
            "mobile": mobile or "",
            "mrd": existing_mrd or "",
            "visit_date": visit_date,
            "tests": tests_list,
            "amount": amount,
            "ref_dr": ref_dr,
            "status": status,
            "errors": errors,
        })

        if status in ("ok", "followup"):
            parsed_for_commit.append({
                "name": name,
                "age": age,
                "gender": gender,
                "dob": dob,
                "mobile": mobile,
                "alternate_mobile": _normalise_mobile(row.get("alternate_mobile", "")),
                "email": email,
                "address": row.get("address") or None,
                "city": row.get("city") or None,
                "state": row.get("state") or None,
                "pincode": (row.get("pincode") or "").strip() or None,
                "occupation": row.get("occupation") or None,
                "chief_complaint": row.get("chief_complaint") or None,
                "referral_source": row.get("referral_source") or None,
                "notes": row.get("notes") or None,
                "existing_mrd": existing_mrd,
                # Visit fields (consumed during commit, not stored on patient doc)
                "_visit_date": visit_date,
                "_bill_no": bill_no,
                "_tests": tests_list,
                "_diagnosis": diagnosis,
                "_amount": amount,
                "_ref_dr": ref_dr,
                "_is_followup": is_followup,
            })

    if not rows_out:
        raise HTTPException(400, "CSV had no data rows.")

    import_id = f"imp_{uuid.uuid4().hex[:16]}"
    expires_at = datetime.utcnow() + timedelta(hours=PREVIEW_TTL_HOURS)
    await db.import_jobs.insert_one(serialize_datetime({
        "import_id": import_id,
        "clinic_id": user["clinic_id"],
        "uploaded_by": user["user_id"],
        "filename": file.filename,
        "tally": {"total": len(rows_out), **counts},
        "rows": parsed_for_commit,           # Only OK rows; skip/fail are not persisted to DB.
        "preview_rows": rows_out,            # Full preview stored for audit + UI.
        "status": "preview",
        "created_at": datetime.utcnow(),
        "expires_at": expires_at,
    }))

    return {
        "import_id": import_id,
        "tally": {"total": len(rows_out), **counts},
        "rows": rows_out,
        "expires_at": expires_at.isoformat(),
    }


# ---------- commit ---------------------------------------------------------

@router.post("/patients/commit")
async def commit_patients(
    payload: dict,
    user=Depends(require_roles("clinic_owner", "super_admin")),
    db=Depends(get_db),
):
    """Commits a previously-previewed CSV by import_id.

    Side-effects per row (NEW — rich import flow):
      * Patient: insert if new, else reuse existing (matched by MR.NO or mobile).
      * Appointment: created on visit_date if present (status=completed).
      * Visit note: PatientNote with tests + diagnosis appended.
      * Invoice + Payment: created if amount > 0. Uses bill_no as `external_invoice_no`,
        else auto-generated. Unknown tests auto-create a generic Service entry.
      * Referring doctor: upserted in `referring_doctors` if Ref.Dr column was provided.
      * MRD policy: `mrd_policy` body param ("keep" | "auto"). Default "keep" — uses the
        clinic's own MR.NO from the CSV verbatim. "auto" → falls through to AUDINEXA's
        sequence generator.

    Idempotent on re-call — committed jobs return their original tally without
    re-inserting.
    """
    import_id = (payload or {}).get("import_id")
    mrd_policy = (payload or {}).get("mrd_policy", "keep").strip().lower()
    if mrd_policy not in ("keep", "auto"):
        mrd_policy = "keep"
    if not import_id:
        raise HTTPException(400, "import_id is required")

    job = await db.import_jobs.find_one(
        {"import_id": import_id, "clinic_id": user["clinic_id"]}, {"_id": 0}
    )
    if not job:
        raise HTTPException(404, "Import preview not found or expired. Please re-upload.")
    if job["status"] == "committed":
        return {
            "import_id": import_id,
            "tally": job.get("commit_tally", job.get("tally", {})),
            "already_committed": True,
        }

    clinic = await db.clinics.find_one({"clinic_id": user["clinic_id"]}, {"_id": 0}) or {}
    mrd_prefix = clinic.get("mrd_prefix", "ACS")

    # Pre-load all patients in the clinic for fast follow-up matching (mrd / mobile).
    patient_by_mrd: dict[str, dict] = {}
    patient_by_mobile: dict[str, dict] = {}
    async for p in db.patients.find(
        {"clinic_id": user["clinic_id"]},
        {"_id": 0, "patient_id": 1, "mrd": 1, "mobile": 1, "name": 1},
    ):
        if p.get("mrd"):
            patient_by_mrd[str(p["mrd"]).strip().upper()] = p
        if p.get("mobile"):
            patient_by_mobile[_normalise_mobile(p["mobile"]) or ""] = p

    # Cache services + ref-docs created during this commit so we don't re-create them.
    services_cache: dict[str, dict] = {}        # token → service doc
    referring_cache: dict[str, str] = {}        # name lower → doctor_id

    async def _resolve_service(token: str, fallback_price: float) -> dict:
        """Find or auto-create a Service for an imported test token (e.g. 'PTA')."""
        key = token.upper()
        if key in services_cache:
            return services_cache[key]
        existing = await db.services.find_one(
            {"clinic_id": user["clinic_id"],
             "$or": [{"code": {"$regex": f"^{re.escape(key)}$", "$options": "i"}},
                     {"name": {"$regex": f"^{re.escape(key)}$", "$options": "i"}}]},
            {"_id": 0},
        )
        if existing:
            services_cache[key] = existing
            return existing
        # Auto-create a minimal service so revenue can attribute by test.
        from uuid import uuid4
        svc = {
            "service_id": f"SVC-{str(uuid4())[:8].upper()}",
            "clinic_id": user["clinic_id"],
            "code": key,
            "name": key,
            "category": "Audiology",
            "price": round(fallback_price, 2),
            "gst_rate": 0.0,
            "gst_inclusive": True,
            "is_taxable": False,
            "active": True,
            "auto_created_via": "import",
            "created_at": datetime.utcnow(),
        }
        await db.services.insert_one(serialize_datetime(dict(svc)))
        services_cache[key] = svc
        return svc

    async def _resolve_ref_doctor(name: str) -> Optional[str]:
        """Find or auto-create a referring doctor for the import's Ref.Dr column."""
        clean = name.strip()
        if not clean:
            return None
        key = clean.lower()
        if key in referring_cache:
            return referring_cache[key]
        existing = await db.referring_doctors.find_one(
            {"clinic_id": user["clinic_id"],
             "name": {"$regex": f"^{re.escape(clean)}$", "$options": "i"}},
            {"_id": 0, "doctor_id": 1},
        )
        if existing:
            referring_cache[key] = existing["doctor_id"]
            return existing["doctor_id"]
        from uuid import uuid4
        did = f"DR-{str(uuid4())[:8].upper()}"
        await db.referring_doctors.insert_one(serialize_datetime({
            "doctor_id": did,
            "clinic_id": user["clinic_id"],
            "name": clean,
            "specialty": None,
            "auto_created_via": "import",
            "created_at": datetime.utcnow(),
        }))
        referring_cache[key] = did
        return did

    created = 0
    followups = 0
    appointments_created = 0
    invoices_created = 0
    payments_total = 0.0
    failed = 0
    failure_details: list[dict] = []

    for r in job.get("rows", []):
        try:
            existing_mrd = r.pop("existing_mrd", None)
            visit_date = r.pop("_visit_date", None)
            bill_no = r.pop("_bill_no", None)
            tests_list = r.pop("_tests", []) or []
            diagnosis = r.pop("_diagnosis", None)
            amount = float(r.pop("_amount", 0) or 0)
            ref_dr = r.pop("_ref_dr", None)
            r.pop("_is_followup", None)

            # Resolve patient: existing or new
            existing_pat = None
            if existing_mrd and existing_mrd in patient_by_mrd:
                existing_pat = patient_by_mrd[existing_mrd]
            elif r.get("mobile"):
                existing_pat = patient_by_mobile.get(r["mobile"])

            if existing_pat:
                patient_id = existing_pat["patient_id"]
                patient_mrd = existing_pat.get("mrd") or existing_mrd
                patient_name = existing_pat.get("name") or r["name"]
                followups += 1
            else:
                # Apply MRD policy
                if mrd_policy == "keep" and existing_mrd:
                    mrd_to_assign = existing_mrd
                else:
                    mrd_to_assign = await _next_mrd(db, user["clinic_id"], mrd_prefix)
                # Patient model requires gender + age non-null. Default missing gender to "Other".
                pat_kwargs = {k: v for k, v in r.items() if not k.startswith("_")}
                if not pat_kwargs.get("gender"):
                    pat_kwargs["gender"] = "Other"
                pat_kwargs["referring_physician"] = ref_dr or pat_kwargs.get("chief_complaint") and None
                ref_doctor_id = await _resolve_ref_doctor(ref_dr) if ref_dr else None
                if ref_doctor_id:
                    pat_kwargs["referring_doctor_id"] = ref_doctor_id
                if ref_dr:
                    pat_kwargs["referring_physician"] = ref_dr
                patient_obj = Patient(
                    **pat_kwargs,
                    clinic_id=user["clinic_id"],
                    mrd=mrd_to_assign,
                )
                doc = serialize_datetime(patient_obj.model_dump())
                await db.patients.insert_one(dict(doc))   # copy so Mongo's _id mutation doesn't leak
                patient_id = patient_obj.patient_id
                patient_mrd = mrd_to_assign
                patient_name = patient_obj.name
                # Cache so later rows in the same import treat them as follow-ups.
                if mrd_to_assign:
                    patient_by_mrd[str(mrd_to_assign).strip().upper()] = {
                        "patient_id": patient_id, "mrd": mrd_to_assign,
                        "mobile": r.get("mobile"), "name": patient_name,
                    }
                if r.get("mobile"):
                    patient_by_mobile[r["mobile"]] = {
                        "patient_id": patient_id, "mrd": mrd_to_assign,
                        "mobile": r.get("mobile"), "name": patient_name,
                    }
                created += 1

            # Side effects: appointment / invoice / visit-note, all keyed on visit_date.
            if visit_date or tests_list or diagnosis or amount > 0:
                from uuid import uuid4
                visit_dt = None
                if visit_date:
                    try:
                        visit_dt = datetime.strptime(visit_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                    except ValueError:
                        visit_dt = datetime.utcnow().replace(tzinfo=timezone.utc)
                else:
                    visit_dt = datetime.utcnow().replace(tzinfo=timezone.utc)

                # Appointment
                ref_doctor_id = await _resolve_ref_doctor(ref_dr) if ref_dr else None
                apt_id = f"APT-{str(uuid4())[:10].upper()}"
                await db.appointments.insert_one(serialize_datetime({
                    "appointment_id": apt_id,
                    "clinic_id": user["clinic_id"],
                    "patient_id": patient_id,
                    "patient_name": patient_name,
                    "patient_mobile": r.get("mobile"),
                    "mrd": patient_mrd,
                    "counterparty_type": "patient",
                    "counterparty_name": patient_name,
                    "counterparty_phone": r.get("mobile"),
                    "service": " + ".join(tests_list) if tests_list else None,
                    "category": "diagnostic" if tests_list else "consultation",
                    "priority": "normal",
                    "visit_type": "referral" if ref_dr else "walkin",
                    "recommended_tests": tests_list,
                    "referred_by": ref_dr,
                    "start_at": visit_dt,
                    "end_at": visit_dt + timedelta(minutes=30),
                    "duration_minutes": 30,
                    "status": "completed",
                    "notes": diagnosis,
                    "created_at": datetime.utcnow(),
                    "created_by_user_id": user["user_id"],
                    "imported_via": import_id,
                }))
                appointments_created += 1

                # Visit note (tests + diagnosis go into patient timeline)
                summary_bits = []
                if tests_list:
                    summary_bits.append("Tests: " + ", ".join(tests_list))
                if diagnosis:
                    summary_bits.append("Diagnosis: " + diagnosis)
                if ref_dr:
                    summary_bits.append("Ref: " + ref_dr)
                if amount > 0:
                    summary_bits.append(f"Paid ₹{amount:.0f}")
                if bill_no:
                    summary_bits.append(f"Bill {bill_no}")
                if summary_bits:
                    await db.patient_notes.insert_one(serialize_datetime({
                        "note_id": f"NOTE-{str(uuid4())[:10].upper()}",
                        "patient_id": patient_id,
                        "audiologist": None,
                        "text": " · ".join(summary_bits),
                        "auto": True,
                        "imported_via": import_id,
                        "visit_date": visit_date,
                        "created_at": visit_dt,
                    }))

                # Invoice + Payment if amount > 0
                if amount > 0:
                    # Distribute amount evenly across tests for service-level revenue attribution.
                    tcount = max(1, len(tests_list))
                    per_test_price = round(amount / tcount, 2)
                    inv_lines = []
                    for t in (tests_list or ["CONSULT"]):
                        svc = await _resolve_service(t, per_test_price)
                        inv_lines.append({
                            "line_id": str(uuid4())[:8],
                            "service_id": svc["service_id"],
                            "description": svc["name"],
                            "quantity": 1,
                            "unit_price": per_test_price,
                            "discount_amount": 0.0,
                            "discount_type": "flat",
                            "discount_value": 0.0,
                            "is_taxable": False,
                            "gst_rate": 0.0,
                            "taxable_value": per_test_price,
                            "cgst_amount": 0.0, "sgst_amount": 0.0, "igst_amount": 0.0,
                            "line_total": per_test_price,
                        })
                    # NAV-008 · Policy B — preserve supplied invoice_no
                    # (from `bill_no` CSV column) OR mint an IMP/… canonical
                    # placeholder. THEN check the compound uniqueness rule
                    # BEFORE insertion. On collision we surface a controlled
                    # per-row failure into `failure_details` — the operator
                    # sees the offending row and can fix the source CSV. We
                    # deliberately do NOT silently overwrite or silently
                    # renumber historical imports; renumbering historical
                    # data is a GST-affecting operation gated separately.
                    invoice_no = bill_no or f"IMP/{visit_date or datetime.utcnow().strftime('%Y-%m-%d')}/{str(uuid4())[:6].upper()}"
                    invoice_id = f"INV-{str(uuid4())[:10].upper()}"
                    payment_id = f"PAY-{str(uuid4())[:8].upper()}"
                    # Duplicate check — if a same-clinic invoice already
                    # exists with this number, reject THIS row and continue
                    # with the next.
                    _existing = await db.invoices.find_one(
                        {"clinic_id": user["clinic_id"], "invoice_no": invoice_no},
                        {"_id": 0, "invoice_id": 1},
                    )
                    if _existing:
                        failed += 1
                        failure_details.append({
                            "row": r.get("name"),
                            "error": (
                                f"Invoice number {invoice_no!r} already exists in this clinic "
                                f"(existing invoice_id={_existing.get('invoice_id')!r}). "
                                f"NAV-008 Policy B: historical imports are never silently "
                                f"renumbered; please deduplicate the source CSV or omit the "
                                f"bill_no column to receive an auto-generated IMP/… number."
                            ),
                        })
                        continue
                    inv_doc = {
                        "invoice_id": invoice_id,
                        "clinic_id": user["clinic_id"],
                        "invoice_no": invoice_no,
                        "external_invoice_no": bill_no,            # original clinic bill # (audit trail)
                        "patient_id": patient_id,
                        "patient_name": patient_name,
                        "patient_mobile": r.get("mobile"),
                        "mrd": patient_mrd,
                        "appointment_id": apt_id,
                        "invoice_date": visit_dt,
                        "lines": inv_lines,
                        "subtotal": amount, "discount_total": 0.0,
                        "cgst_total": 0.0, "sgst_total": 0.0, "igst_total": 0.0, "tax_total": 0.0,
                        "grand_total": amount, "rounded_total": round(amount), "round_off": round(amount) - amount,
                        "paid_total": amount, "due_total": 0.0,
                        "status": "paid",
                        "payments": [{
                            "payment_id": payment_id,
                            "clinic_id": user["clinic_id"],
                            "invoice_id": invoice_id,
                            "method": "cash",
                            "amount": amount,
                            "reference": bill_no,
                            "paid_at": visit_dt,
                            "received_by_user_id": user["user_id"],
                            "notes": "Imported from CSV",
                        }],
                        "notes": diagnosis,
                        "created_at": datetime.utcnow(),
                        "created_by_user_id": user["user_id"],
                        "imported_via": import_id,
                    }
                    try:
                        await db.invoices.insert_one(serialize_datetime(dict(inv_doc)))
                    except Exception as _dup:
                        # If the compound unique index rejected this row
                        # despite our pre-check (rare race), surface it
                        # as a per-row failure rather than aborting the
                        # whole import.
                        if "E11000" in str(_dup) or "duplicate" in str(_dup).lower():
                            failed += 1
                            failure_details.append({
                                "row": r.get("name"),
                                "error": (
                                    f"Invoice number {invoice_no!r} collided at insert "
                                    f"time (concurrent race). Row skipped. Retry the "
                                    f"import for this row to receive a fresh IMP/… number."
                                ),
                            })
                            continue
                        raise
                    # Also write a top-level payment row so revenue aggregation picks it up.
                    await db.payments.insert_one(serialize_datetime({
                        "payment_id": payment_id,
                        "clinic_id": user["clinic_id"],
                        "invoice_id": invoice_id,
                        "method": "cash",
                        "amount": amount,
                        "reference": bill_no,
                        "paid_at": visit_dt,
                        "received_by_user_id": user["user_id"],
                        "notes": "Imported from CSV",
                        "imported_via": import_id,
                        "tests": tests_list,
                        "referring_doctor_id": ref_doctor_id,
                        "referring_doctor_name": ref_dr,
                        "patient_id": patient_id,
                        "visit_date": visit_date,
                    }))
                    invoices_created += 1
                    payments_total += amount

        except Exception as exc:
            failed += 1
            failure_details.append({"row": r.get("name"), "error": str(exc)})

    commit_tally = {
        "created": created,
        "followups": followups,
        "appointments": appointments_created,
        "invoices": invoices_created,
        "revenue": round(payments_total, 2),
        "failed": failed,
        "skipped": job.get("tally", {}).get("will_skip", 0),
    }

    await db.import_jobs.update_one(
        {"import_id": import_id},
        {"$set": serialize_datetime({
            "status": "committed",
            "committed_at": datetime.utcnow(),
            "committed_by": user["user_id"],
            "mrd_policy": mrd_policy,
            "commit_tally": commit_tally,
            "failure_details": failure_details,
        })},
    )
    await db.activity_logs.insert_one(serialize_datetime({
        "clinic_id": user["clinic_id"],
        "user_id": user["user_id"],
        "action": "patient.bulk_import",
        "import_id": import_id,
        "tally": commit_tally,
        "at": datetime.utcnow(),
    }))

    return {
        "import_id": import_id,
        "tally": commit_tally,
        "failure_details": failure_details,
        "already_committed": False,
    }


# ---------- recent imports (audit panel) -----------------------------------

@router.get("/patients/recent")
async def list_recent_imports(
    user=Depends(require_roles("clinic_owner", "super_admin")),
    db=Depends(get_db),
):
    """Last 20 import jobs for the clinic — drives the 'history' strip in the UI."""
    cursor = db.import_jobs.find(
        {"clinic_id": user["clinic_id"]},
        {"_id": 0, "import_id": 1, "filename": 1, "tally": 1, "commit_tally": 1,
         "status": 1, "created_at": 1, "committed_at": 1},
    ).sort("created_at", -1).limit(20)
    return [doc async for doc in cursor]
